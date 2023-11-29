from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

Workbook = load_workbook("../Resources/Datadriven_TestData.xlsx") # To open an existing workbook

def create_new(sheetname):
    wb= Workbook()
    ws= wb.active()
    wb.save("../Resources/Created_Test.xlsx")

    # when a worksheet is created in memory, it won't contain cells. cells will be created when they are accessed
    ws.title = sheetname
    ws2 = wb.create_sheet("Mysheet",-1)
    ws2['D18']=250   # Cell creation
    print(ws2['D18'].value) # printing a cell value



def fetch_max_rows(sheetname):
    Sh = Workbook[sheetname]
    return Sh.max_row


def fetch_max_col(sheetname):
    Sh = Workbook[sheetname]
    return Sh.max_column


def fetch_cellData(sheetname, row, column):
    Sh = Workbook[sheetname]
    Cell = Sh.cell(int(row), int(column))
    return Cell.value






